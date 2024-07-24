
import streamlit as st
import os
import zipfile
from io import BytesIO
from PIL import ImageOps, ImageChops
from PIL import Image as PILImage
import shutil
import pytesseract

import cv2
import numpy as np
from pytesseract import Output

import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image as openpyxlImage

#st.layout(width="80%")



def trim_image(image, border=10):
    # 使用 getbbox 來裁剪圖片周圍的空白部分
    bg = PILImage.new(PILImage.mode, PILImage.size, PILImage.getpixel((0, 0)))
    diff = ImageOps.invert(ImageChops.difference(image, bg))
    bbox = diff.getbbox()
    if bbox:
        # 裁剪圖片
        image = PILImage.crop(bbox)
        # 添加邊距
        return ImageOps.expand(image, border=border, fill=image.getpixel((0, 0)))
    return image

def clear_ocr_files_directory(directory):
    if os.path.exists(directory):
        shutil.rmtree(directory)
    os.makedirs(directory)


def save_uploaded_file(uploaded_file, directory):
    try:
        # 清空 ocr_files 目錄
        clear_ocr_files_directory(directory)
        
        # 保存上傳的 ZIP 文件到臨時緩存中
        with BytesIO(uploaded_file.read()) as zf:
            with zipfile.ZipFile(zf, 'r') as zip_ref:
                # 指定解壓縮時使用的編碼方式
                for zip_info in zip_ref.infolist():
                    zip_info.filename = zip_info.filename.encode('cp437').decode('big5')  # 修改這一行，根據實際情況調整編碼
                    zip_ref.extract(zip_info, directory)
        return True
    except Exception as e:
        st.error(f"出現錯誤: {e}")
        return False

@st.cache_data
def BSMI(item):

   
    with open(item, 'rb') as f:
        image_cv2 = cv2.imdecode(np.frombuffer(f.read(), np.uint8), cv2.IMREAD_COLOR)    
   
    assert image_cv2 is not None, "file could not be read, check with os.path.exists()"
    
    gray = cv2.cvtColor(image_cv2, cv2.COLOR_BGR2GRAY)
    
    img2 = gray.copy()
    
    
    template = cv2.imread('BSMIPNG.png', cv2.IMREAD_GRAYSCALE)
    assert template is not None, "file could not be read, check with os.path.exists()"
    
    #w, h = template.shape[::-1]
    # All the 6 methods for comparison in a list
    
    methods = ['TM_CCOEFF', 'TM_CCOEFF_NORMED', 'TM_CCORR',
     'TM_CCORR_NORMED', 'TM_SQDIFF', 'TM_SQDIFF_NORMED']

    meth = 'TM_CCOEFF_NORMED'
    img = img2.copy()
    method = getattr(cv2, meth)

    min_max_best = -99
    scale_best = -1
    top_left_best = []
    #bsmi_find = False
    w_best = -1
    h_best = -1

    for scale in np.arange(0.1, 1.0, 0.01):
        resized_template = cv2.resize(template, (0, 0), fx=scale, fy=scale)
        
        w, h = resized_template.shape[::-1]

        # Apply template Matching
        res = cv2.matchTemplate(img,resized_template,method)
        min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(res)

        top_left = max_loc
        min_max = max_val

        if(min_max_best > 0.7 and min_max < min_max_best):
            break
        
        if(min_max > min_max_best):
            min_max_best = min_max
            scale_best = scale
            top_left_best = top_left
            w_best = w
            h_best = h
            
            
            


    bottom_right = (top_left_best[0] + w_best, top_left_best[1] + h_best)
    cv2.rectangle(image_cv2, top_left_best, bottom_right, (255, 0, 0), 2)
    
    return image_cv2,min_max_best,scale_best,top_left_best


@st.cache_data
def batch_ocr(items):

        
        
        # 取得當前日期和時間
        now = datetime.datetime.now()
        formatted_time = now.strftime("%Y%m%d%H%M%S")

        # 建立文件名
        filename = f"output_{formatted_time}.xlsx"

        # 創建新工作簿
        wb = Workbook()
        ws = wb.active

        # 在 B1 儲存格存儲文件名內容
        ws['A1'] = "序號"
        ws['B1'] = "檔案名稱"
        ws.column_dimensions['B'].width = 40
        
        ws['C1'] = "OCR結果"
        ws.column_dimensions['C'].width = 60        


        ws['D1'] = "最佳係數(TM_CCOEFF_NORMED)"
        ws.column_dimensions['D'].width = 30        
        
        
        ws['E1'] = "最佳尺度(scale)"
        ws.column_dimensions['E'].width = 20        
        
        ws['F1'] = "BSMI識別"
        ws.column_dimensions['F'].width = 20        
        
        ws['G1'] = "BSMI座標"
        ws.column_dimensions['G'].width = 20
        
        # ws['H1'] = "BSMI抓圖"
        
        ws['I1'] = "原圖檔"
        ws['J1'] = "識別圖檔"
        ws.column_dimensions['J'].width = 50   
        

        row_num = 2
        number = 0
        
        alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # 顯示文件和子目錄
        for item in items:
            if os.path.isfile(item) and item.lower().endswith('.png'):
            
                ws.row_dimensions[row_num].height = 350
                
                #ws.row_dimensions[row_num].alignment = alignment
            
                number = number + 1
                ws['A'+str(row_num)] = number
                ws['B'+str(row_num)] = item

                image = PILImage.open(item)
                text_psm6 = pytesseract.image_to_string(image, lang='chi_tra', config='--psm 6')
                ws['C'+str(row_num)] = text_psm6

                #BSMI start
                image_cv2,min_max_best,scale_best,top_left_best = BSMI(item)

                ws['D'+str(row_num)] = min_max_best
                ws['E'+str(row_num)] = scale_best

                if(min_max_best >= 0.7):
                    ws['F'+str(row_num)] = "存在"
                    ws['G'+str(row_num)] = str(top_left_best)
                else:
                    ws['F'+str(row_num)] = "不存在"
                   
                
                
                
                # 將 OpenCV 圖片轉換為 Pillow 圖片
                #image_pil = PILImage.fromarray(cv2.cvtColor(image_cv2, cv2.COLOR_BGR2RGB))
                
                image_pil = PILImage.fromarray(image_cv2)
                
                
                desired_width_px = 300
                aspect_ratio = image_pil.height / image_pil.width
                new_height_px = int(desired_width_px * aspect_ratio)
                
                # 調整圖片大小
                resized_image_pil = image_pil.resize((desired_width_px, new_height_px))


                # 將 PIL Image 儲存到 BytesIO 物件
                image_stream = BytesIO()
                resized_image_pil.save(image_stream, format='PNG')
                image_stream.seek(0)

                # 將影像新增到 J5 儲存格
                img = openpyxlImage(image_stream)
                ws.add_image(img, 'J'+str(row_num))                

                #row_num = row_num + 1
                row_num = row_num + 1
        

        # 遍歷A到J欄，1到100列的所有儲存格並應用對齊方式
        for row in ws.iter_rows(min_row=2, max_row=row_num, min_col=1, max_col=20):
            for cell in row:
                cell.alignment = alignment
        
        #ws['B2'] = filename

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # 儲存工作簿
        #wb.save(filename)
        #st.write(f"檔案已儲存為: {filename}") 
        
        
        st.write("批次處理完成")
        
        return filename,excel_file

  




def list_files_in_directory(directory):
    # 遞迴列出目錄中的所有文件和子目錄
    files_and_dirs = []
    for root, dirs, files in os.walk(directory):
        for name in files:
            if name.lower().endswith('.png'):
                files_and_dirs.append(os.path.join(root, name))
        #for name in dirs:
        #    files_and_dirs.append(os.path.join(root, name))
    
	#return files_and_dirs
	

	# 列出目錄中的所有文件和子目錄
    items = files_and_dirs

    if not items:
        st.write("目錄中沒有文件或子目錄。")
        return

    if st.button(f"批次處理", key="batch"):
        
        filename,excel_file = batch_ocr(items)

        # 顯示下載連結
        st.download_button(
            label=f"下載 {filename}",
            data=excel_file,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )      
        
        return
        
        
        image = PILImage.open(item)
        # image2 = trim_image(image)

        #text = pytesseract.image_to_string(image, lang='chi_tra')
        
        #text_psm3 = pytesseract.image_to_string(image, lang='chi_tra', config='--psm 3')
        #st.write(text)
        #cols[0].markdown('*tesseract --psm 3:*')
        #cols[0].write(text_psm3)

        text_psm6 = pytesseract.image_to_string(image, lang='chi_tra', config='--psm 6')
        
        cols[0].markdown('*OCR 結果:*')
        cols[0].write(text_psm6)

        # 在兩欄之間插入一條分隔線
        st.markdown("""
        <style>
        div[data-testid="column"] {
            border-right: 1px solid #e6e6e6;
            padding-right: 20px;
        }
        </style>
        """, unsafe_allow_html=True)
        
        #BSMI start
        
        image_cv2,min_max_best,scale_best,top_left_best = BSMI(item)
        
        BSMI_result = '不存在'
        if(min_max_best >= 0.7):
            BSMI_result = f'存在; 座標:{top_left_best}'
        
        
        cols[1].markdown('*BSMI識別結果:*')
        
        cols[1].write(f'最佳係數(TM_CCOEFF_NORMED):{min_max_best:.4f}')
        
        cols[1].write(f'最佳尺度(scale):{scale_best:.4f}')
        
        cols[1].write(f'BSMI圖形:{BSMI_result}')
        
        #BSMI end

        st.divider()

        st.markdown('*原圖檔:*')
        st.image(image, caption=os.path.basename(item))
        
        st.markdown('*識別圖檔:*')
        st.image(image_cv2, caption=os.path.basename(item))
        
        
        
        #cols[0].image(image2, caption=os.path.basename(item))
    

    st.write("文件列表：")

    # 顯示文件和子目錄
    for item in items:
        if os.path.isfile(item) and item.lower().endswith('.png'):
            #cols = st.columns(2)
            st.write(f"{item}")    
			
         
            
            
	
	

def main():
    #st.title("OCR Files Browser")

    directory = 'batch_files'  # 這是你要顯示的目錄


    # 創建文件上傳控件
    
    
    
	#st.markdown("### 上傳並解壓縮 ZIP 檔案")
    
	#uploaded_file = st.file_uploader("選擇一個 ZIP 檔案", type=["zip"])

	#if uploaded_file is not None:
	#	if save_uploaded_file(uploaded_file, directory):
	#		st.success("文件上傳並解壓縮成功！")
     

            # 列出解壓縮後的文件
            # st.write("解壓縮後的文件列表：")
            # 列出目錄中的所有文件和子目錄


	#	else:
	#		st.error("文件上傳或解壓縮失敗。")
    
			
    list_files_in_directory(directory)

main()

